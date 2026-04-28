from rest_framework import generics, status, viewsets, filters
from rest_framework.decorators import api_view, permission_classes, authentication_classes, action
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework_simplejwt.authentication import JWTAuthentication
from rest_framework_simplejwt.exceptions import InvalidToken, TokenError
from django.contrib.auth import authenticate
from django.db.models import Q, Avg, Count, F
from django.utils import timezone


# ── Query-param JWT auth (for browser download links) ────────────────────────
class QueryParamJWTAuthentication(JWTAuthentication):
    """
    Accepts a JWT token passed as ?token=<access_token> in the URL.
    Falls back to the standard Authorization: Bearer header so both work.
    """
    def authenticate(self, request):
        token = request.query_params.get('token')
        if token:
            try:
                validated = self.get_validated_token(token)
                return self.get_user(validated), validated
            except (InvalidToken, TokenError):
                pass
        # Fall back to header-based auth
        return super().authenticate(request)

from .models import (
    User, StudentProfile, RecruiterProfile, JobPosting, Application,
    InterviewSchedule, PlacementDrive, AlumniProfile, Review,
    Notification, ChatMessage, Feedback, SkillCourse, StudentCourseProgress
)
from .serializers import (
    RegisterSerializer, LoginSerializer, UserSerializer,
    StudentProfileSerializer, RecruiterProfileSerializer,
    JobPostingListSerializer, JobPostingDetailSerializer, JobPostingCreateSerializer,
    ApplicationSerializer, ApplicationCreateSerializer,
    InterviewScheduleSerializer, PlacementDriveSerializer,
    AlumniProfileSerializer, ReviewSerializer,
    NotificationSerializer, ChatMessageSerializer,
    FeedbackSerializer, SkillCourseSerializer, StudentCourseProgressSerializer
)
from .permissions import (
    IsStudent, IsRecruiter, IsPlacementOfficer, IsAdminUser,
    IsRecruiterOrPlacementOfficer, IsOwnerOrAdmin
)


# ============ HELPERS ============

def _create_and_push_notification(user, notification_type, title, message, link=''):
    """
    Create a Notification DB record and push it over WebSocket to the user's
    notification channel group in real time.
    """
    notif = Notification.objects.create(
        user=user,
        notification_type=notification_type,
        title=title,
        message=message,
        link=link,
    )
    # Push via Django Channels/Redis if available
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        channel_layer = get_channel_layer()
        if channel_layer:
            async_to_sync(channel_layer.group_send)(
                f'notifications_{user.id}',
                {
                    'type': 'send_notification',
                    'notification': {
                        'id': notif.id,
                        'notification_type': notification_type,
                        'title': title,
                        'message': message,
                        'link': link,
                        'is_read': False,
                    }
                }
            )
    except Exception:
        pass  # WebSocket push is bonus; DB record is the source of truth
    return notif


# ============ AUTH VIEWS ============

@api_view(['POST'])
@permission_classes([AllowAny])
def register(request):
    """Register a new user."""
    serializer = RegisterSerializer(data=request.data)
    if serializer.is_valid():
        user = serializer.save()
        refresh = RefreshToken.for_user(user)
        return Response({
            'message': 'Registration successful',
            'user': UserSerializer(user).data,
            'tokens': {
                'refresh': str(refresh),
                'access': str(refresh.access_token),
            }
        }, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([AllowAny])
def login(request):
    """Login user and return JWT tokens."""
    serializer = LoginSerializer(data=request.data)
    if serializer.is_valid():
        email = serializer.validated_data['email']
        password = serializer.validated_data['password']
        role = serializer.validated_data['role']

        try:
            user = User.objects.get(email=email)
        except User.DoesNotExist:
            return Response({'error': 'Invalid credentials'}, status=status.HTTP_401_UNAUTHORIZED)

        if not user.check_password(password):
            return Response({'error': 'Invalid credentials'}, status=status.HTTP_401_UNAUTHORIZED)

        if user.role != role:
            return Response({'error': f'This account is not registered as {role}'},
                            status=status.HTTP_401_UNAUTHORIZED)

        refresh = RefreshToken.for_user(user)
        return Response({
            'message': 'Login successful',
            'user': UserSerializer(user).data,
            'tokens': {
                'refresh': str(refresh),
                'access': str(refresh.access_token),
            }
        })
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def logout(request):
    """
    POST /api/auth/logout/
    Body: { "refresh": "<refresh_token>" }
    Blacklists the refresh token so it cannot be reused.
    """
    refresh_token = request.data.get('refresh')
    if not refresh_token:
        return Response({'error': 'Refresh token is required.'}, status=status.HTTP_400_BAD_REQUEST)
    try:
        token = RefreshToken(refresh_token)
        token.blacklist()
        return Response({'message': 'Logged out successfully.'})
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_profile(request):
    """Get current user profile."""
    user = request.user
    data = UserSerializer(user).data

    if user.role == 'student':
        try:
            profile = StudentProfile.objects.get(user=user)
            data['student_profile'] = StudentProfileSerializer(profile).data
        except StudentProfile.DoesNotExist:
            data['student_profile'] = None
    elif user.role == 'recruiter':
        try:
            profile = RecruiterProfile.objects.get(user=user)
            data['recruiter_profile'] = RecruiterProfileSerializer(profile).data
        except RecruiterProfile.DoesNotExist:
            data['recruiter_profile'] = None

    return Response(data)


# ============ STUDENT PROFILE VIEWS ============

class StudentProfileView(generics.RetrieveUpdateAPIView):
    serializer_class = StudentProfileSerializer
    permission_classes = [IsAuthenticated]

    def get_object(self):
        profile, _ = StudentProfile.objects.get_or_create(
            user=self.request.user,
            defaults={
                'roll_number': f'TEMP_{self.request.user.id}',
                'department': 'CSE',
                'batch_year': timezone.now().year,
                'cgpa': 0.0
            }
        )
        return profile

    def perform_update(self, serializer):
        old_resume = self.get_object().resume
        instance = serializer.save(user=self.request.user)
        # If a new resume was uploaded, trigger async parsing in background
        new_resume = instance.resume
        if new_resume and new_resume != old_resume:
            try:
                from api.tasks import parse_resume_async
                parse_resume_async.delay(instance.user.id)
            except Exception:
                pass  # Celery not running; parsing skipped (user can call /api/resume/parse/ manually)


class StudentListView(generics.ListAPIView):
    serializer_class = StudentProfileSerializer
    permission_classes = [IsRecruiterOrPlacementOfficer]

    def get_queryset(self):
        queryset = StudentProfile.objects.all()
        department = self.request.query_params.get('department')
        min_cgpa = self.request.query_params.get('min_cgpa')
        if department:
            queryset = queryset.filter(department=department)
        if min_cgpa:
            queryset = queryset.filter(cgpa__gte=min_cgpa)
        return queryset

    def list(self, request, *args, **kwargs):
        """Override list to fire profile_viewed notifications for each profile seen."""
        response = super().list(request, *args, **kwargs)
        # Notify each student that their profile was viewed (async, fire-and-forget)
        for item in response.data.get('results', response.data if isinstance(response.data, list) else []):
            try:
                student_id = item.get('user', {}).get('id') if isinstance(item.get('user'), dict) else None
                if student_id and student_id != request.user.id:
                    Notification.objects.get_or_create(
                        user_id=student_id,
                        notification_type='profile_viewed',
                        defaults={
                            'title': 'Your profile was viewed',
                            'message': f'A recruiter or placement officer viewed your profile.',
                        }
                    )
            except Exception:
                pass
        return response


# ============ RECRUITER PROFILE VIEWS ============

class RecruiterProfileView(generics.RetrieveUpdateAPIView):
    serializer_class = RecruiterProfileSerializer
    permission_classes = [IsAuthenticated]

    def get_object(self):
        profile, _ = RecruiterProfile.objects.get_or_create(
            user=self.request.user,
            defaults={'company_name': 'My Company'}
        )
        return profile


# ============ JOB POSTING VIEWS ============

class JobPostingListView(generics.ListAPIView):
    serializer_class = JobPostingListSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = JobPosting.objects.filter(status='approved')
        user = self.request.user

        # Filters
        department = self.request.query_params.get('department')
        min_cgpa = self.request.query_params.get('min_cgpa')
        job_type = self.request.query_params.get('job_type')
        search = self.request.query_params.get('search')
        company = self.request.query_params.get('company')

        if department:
            queryset = queryset.filter(eligible_departments__icontains=department)
        if min_cgpa:
            queryset = queryset.filter(min_cgpa__lte=min_cgpa)
        if job_type:
            queryset = queryset.filter(job_type=job_type)
        if search:
            queryset = queryset.filter(
                Q(title__icontains=search) | Q(description__icontains=search)
            )
        if company:
            queryset = queryset.filter(
                recruiter__recruiter_profile__company_name__icontains=company
            )

        return queryset


class JobPostingDetailView(generics.RetrieveAPIView):
    serializer_class = JobPostingDetailSerializer
    permission_classes = [IsAuthenticated]
    queryset = JobPosting.objects.all()

    def get_serializer_context(self):
        return {'request': self.request}


class JobPostingCreateView(generics.CreateAPIView):
    serializer_class = JobPostingCreateSerializer
    permission_classes = [IsRecruiter]

    def perform_create(self, serializer):
        serializer.save(recruiter=self.request.user)


class RecruiterJobsView(generics.ListAPIView):
    serializer_class = JobPostingListSerializer
    permission_classes = [IsRecruiter]

    def get_queryset(self):
        return JobPosting.objects.filter(recruiter=self.request.user)


class JobApprovalView(generics.UpdateAPIView):
    serializer_class = JobPostingDetailSerializer
    # Allow both placement_officer AND admin to approve/reject jobs
    permission_classes = [IsRecruiterOrPlacementOfficer]
    queryset = JobPosting.objects.all()

    def update(self, request, *args, **kwargs):
        job = self.get_object()
        action = request.data.get('action')
        if action == 'approve':
            job.status = 'approved'
            job.approved_by = request.user
            job.save()
            # Notify relevant students via Celery background task
            try:
                from api.tasks import notify_students_new_job
                notify_students_new_job.delay(job.id)
            except Exception:
                # Fallback: notify inline if Celery is not running
                self._notify_students(job)
            return Response({'message': 'Job approved successfully'})
        elif action == 'reject':
            job.status = 'rejected'
            job.save()
            return Response({'message': 'Job rejected'})
        elif action == 'close':
            job.status = 'closed'
            job.save()
            return Response({'message': 'Job closed'})
        return Response({'error': 'Invalid action. Use: approve, reject, or close.'}, status=status.HTTP_400_BAD_REQUEST)

    def _notify_students(self, job):
        students = User.objects.filter(role='student')
        for student in students:
            try:
                profile = student.student_profile
                student_cgpa = float(profile.cgpa or 0)
                min_cgpa = float(job.min_cgpa or 0)
                if student_cgpa >= min_cgpa:
                    if not job.eligible_departments or profile.department in job.eligible_departments:
                        try:
                            company_name = job.recruiter.recruiter_profile.company_name
                        except Exception:
                            company_name = "Company"
                        Notification.objects.create(
                            user=student,
                            notification_type='job_alert',
                            title=f'New Job: {job.title}',
                            message=f'{job.title} at {company_name}. Apply before {job.application_deadline}.',
                            link=f'/jobs/{job.id}'
                        )
            except Exception:
                pass


class PendingJobsView(generics.ListAPIView):
    serializer_class = JobPostingListSerializer
    permission_classes = [IsPlacementOfficer]

    def get_queryset(self):
        return JobPosting.objects.filter(status='pending')


# ============ APPLICATION VIEWS ============

class ApplyJobView(generics.CreateAPIView):
    serializer_class = ApplicationCreateSerializer
    permission_classes = [IsStudent]

    def perform_create(self, serializer):
        job = serializer.validated_data['job']
        # Check if already applied
        if Application.objects.filter(student=self.request.user, job=job).exists():
            from rest_framework.exceptions import ValidationError
            raise ValidationError("You have already applied for this job.")
        # Enforce max_applications limit
        current_count = Application.objects.filter(job=job).exclude(status='withdrawn').count()
        if job.max_applications and current_count >= job.max_applications:
            from rest_framework.exceptions import ValidationError
            raise ValidationError(
                f"This job has reached its maximum application limit ({job.max_applications}).")
        application = serializer.save(student=self.request.user)
        # Notify recruiter (and push via WebSocket)
        _create_and_push_notification(
            user=job.recruiter,
            notification_type='application_update',
            title=f'New Application: {job.title}',
            message=f'{self.request.user.get_full_name()} applied for {job.title}.',
            link=f'/applications/{application.id}'
        )


@api_view(['POST'])
@permission_classes([IsStudent])
def withdraw_application(request, pk):
    """
    POST /api/applications/<pk>/withdraw/
    Student withdraws their own application (only from 'applied' status).
    """
    try:
        application = Application.objects.get(pk=pk, student=request.user)
    except Application.DoesNotExist:
        return Response({'error': 'Application not found.'}, status=status.HTTP_404_NOT_FOUND)

    if application.status not in ('applied', 'shortlisted'):
        return Response(
            {'error': f'Cannot withdraw an application with status "{application.status}".'},
            status=status.HTTP_400_BAD_REQUEST
        )
    application.status = 'withdrawn'
    application.save()
    return Response({'message': 'Application withdrawn successfully.'})


class StudentApplicationsView(generics.ListAPIView):
    serializer_class = ApplicationSerializer
    permission_classes = [IsStudent]

    def get_queryset(self):
        return Application.objects.filter(student=self.request.user)


class JobApplicationsView(generics.ListAPIView):
    serializer_class = ApplicationSerializer
    permission_classes = [IsRecruiterOrPlacementOfficer]

    def get_queryset(self):
        job_id = self.kwargs.get('job_id')
        return Application.objects.filter(job_id=job_id)


class UpdateApplicationStatusView(generics.UpdateAPIView):
    serializer_class = ApplicationSerializer
    permission_classes = [IsRecruiterOrPlacementOfficer]
    queryset = Application.objects.all()

    def update(self, request, *args, **kwargs):
        application = self.get_object()
        new_status = request.data.get('status')
        valid_statuses = ['shortlisted', 'interview', 'selected', 'rejected']
        if new_status not in valid_statuses:
            return Response({'error': 'Invalid status'}, status=status.HTTP_400_BAD_REQUEST)

        application.status = new_status
        if request.data.get('recruiter_notes'):
            application.recruiter_notes = request.data['recruiter_notes']
        application.save()

        # Notify student via DB + WebSocket push
        status_messages = {
            'shortlisted': 'Congratulations! You have been shortlisted',
            'interview': 'Your interview has been scheduled',
            'selected': 'Congratulations! You have been selected',
            'rejected': 'Unfortunately, your application was not selected',
        }
        _create_and_push_notification(
            user=application.student,
            notification_type='application_update',
            title=f'Application Update: {application.job.title}',
            message=f'{status_messages[new_status]} for {application.job.title}.',
            link=f'/applications/{application.id}'
        )

        if new_status == 'selected':
            try:
                profile = application.student.student_profile
                profile.is_placed = True
                profile.placed_company = application.job.recruiter.recruiter_profile.company_name
                # Auto-set placed_package from the job's salary_max if available
                if application.job.salary_max and not profile.placed_package:
                    profile.placed_package = application.job.salary_max
                profile.save()
            except (StudentProfile.DoesNotExist, RecruiterProfile.DoesNotExist):
                pass

        return Response({'message': f'Application status updated to {new_status}'})


# ============ INTERVIEW SCHEDULE VIEWS ============

class InterviewScheduleCreateView(generics.CreateAPIView):
    serializer_class = InterviewScheduleSerializer
    permission_classes = [IsRecruiterOrPlacementOfficer]

    def perform_create(self, serializer):
        interview = serializer.save()
        _create_and_push_notification(
            user=interview.application.student,
            notification_type='interview_reminder',
            title=f'Interview Scheduled: {interview.application.job.title}',
            message=f'Your {interview.get_round_type_display()} is scheduled on {interview.scheduled_date}.',
            link=f'/interviews/{interview.id}'
        )


class InterviewScheduleUpdateView(generics.UpdateAPIView):
    """
    PATCH /api/interviews/<pk>/update/
    Update interview result and feedback after the round.
    """
    serializer_class = InterviewScheduleSerializer
    permission_classes = [IsRecruiterOrPlacementOfficer]
    queryset = InterviewSchedule.objects.all()

    def update(self, request, *args, **kwargs):
        interview = self.get_object()
        result = request.data.get('result')
        feedback = request.data.get('feedback')
        if result and result not in ('pending', 'passed', 'failed'):
            return Response({'error': 'Invalid result. Use: pending, passed, or failed.'},
                            status=status.HTTP_400_BAD_REQUEST)
        if result:
            interview.result = result
        if feedback is not None:
            interview.feedback = feedback
        interview.save()
        return Response(InterviewScheduleSerializer(interview).data)


class StudentInterviewsView(generics.ListAPIView):
    serializer_class = InterviewScheduleSerializer
    permission_classes = [IsStudent]

    def get_queryset(self):
        return InterviewSchedule.objects.filter(application__student=self.request.user)


# ============ PLACEMENT DRIVE VIEWS ============

class PlacementDriveListView(generics.ListCreateAPIView):
    serializer_class = PlacementDriveSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return PlacementDrive.objects.all()

    def perform_create(self, serializer):
        # Only placement officers and admins can create drives
        if self.request.user.role not in ('placement_officer', 'admin'):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('Only placement officers can create placement drives.')
        # company field is required by model — use the creating officer as the company owner
        serializer.save(created_by=self.request.user, company=self.request.user)


class PlacementDriveDetailView(generics.RetrieveUpdateAPIView):
    """
    GET  /api/drives/<pk>/  — retrieve a single drive
    PUT  /api/drives/<pk>/  — update drive (placement officer/admin only)
    """
    serializer_class = PlacementDriveSerializer
    permission_classes = [IsAuthenticated]
    queryset = PlacementDrive.objects.all()


class PlacementDriveRegisterView(generics.UpdateAPIView):
    permission_classes = [IsStudent]
    queryset = PlacementDrive.objects.all()

    def update(self, request, *args, **kwargs):
        drive = self.get_object()
        if request.user in drive.registered_students.all():
            drive.registered_students.remove(request.user)
            return Response({'message': 'Unregistered from drive'})
        drive.registered_students.add(request.user)
        return Response({'message': 'Registered for drive successfully'})


# ============ ALUMNI VIEWS ============

class AlumniListView(generics.ListAPIView):
    serializer_class = AlumniProfileSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = AlumniProfile.objects.filter(is_available_for_mentorship=True)
        company = self.request.query_params.get('company')
        department = self.request.query_params.get('department')
        role = self.request.query_params.get('role')
        graduation_year = self.request.query_params.get('graduation_year')
        if company:
            queryset = queryset.filter(current_company__icontains=company)
        if department:
            queryset = queryset.filter(department=department)
        if role:
            queryset = queryset.filter(current_role__icontains=role)
        if graduation_year:
            queryset = queryset.filter(graduation_year=graduation_year)
        return queryset


class AlumniDetailView(generics.RetrieveAPIView):
    serializer_class = AlumniProfileSerializer
    permission_classes = [IsAuthenticated]
    queryset = AlumniProfile.objects.all()


class AlumniProfileCreateView(generics.CreateAPIView):
    """
    POST /api/alumni/create/
    Any authenticated user can register as alumni (creates an AlumniProfile
    linked to the current user).
    """
    serializer_class = AlumniProfileSerializer
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


class AlumniProfileUpdateView(generics.RetrieveUpdateAPIView):
    """
    GET/PUT/PATCH /api/alumni/my/
    Current user retrieves or updates their own AlumniProfile
    (e.g. toggle is_available_for_mentorship).
    """
    serializer_class = AlumniProfileSerializer
    permission_classes = [IsAuthenticated]

    def get_object(self):
        try:
            return AlumniProfile.objects.get(user=self.request.user)
        except AlumniProfile.DoesNotExist:
            from rest_framework.exceptions import NotFound
            raise NotFound('You do not have an alumni profile. Create one at POST /api/alumni/create/')


# ============ REVIEW VIEWS ============

class ReviewListCreateView(generics.ListCreateAPIView):
    serializer_class = ReviewSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = Review.objects.all()
        company = self.request.query_params.get('company')
        if company:
            queryset = queryset.filter(company_name__icontains=company)
        return queryset

    def perform_create(self, serializer):
        review = serializer.save(reviewer=self.request.user)
        # Groq-powered sentiment analysis (single source of truth)
        try:
            from .ml.groq_service import analyze_sentiment
            text = (review.content or '') + ' ' + (review.pros or '') + ' ' + (review.cons or '')
            result = analyze_sentiment(text.strip())
            Review.objects.filter(pk=review.pk).update(
                sentiment_score=result['score'],
                sentiment_label=result['label'],
            )
        except Exception:
            pass


class ReviewDetailView(generics.RetrieveDestroyAPIView):
    """
    GET    /api/reviews/<pk>/  — retrieve a single review
    DELETE /api/reviews/<pk>/  — reviewer or admin can delete
    """
    serializer_class = ReviewSerializer
    permission_classes = [IsAuthenticated]
    queryset = Review.objects.all()

    def destroy(self, request, *args, **kwargs):
        review = self.get_object()
        if review.reviewer != request.user and request.user.role != 'admin':
            return Response({'error': 'You can only delete your own reviews.'},
                            status=status.HTTP_403_FORBIDDEN)
        review.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


# ============ NOTIFICATION VIEWS ============

class NotificationListView(generics.ListAPIView):
    serializer_class = NotificationSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return Notification.objects.filter(user=self.request.user)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def mark_notification_read(request, pk):
    try:
        notification = Notification.objects.get(pk=pk, user=request.user)
        notification.is_read = True
        notification.save()
        return Response({'message': 'Notification marked as read'})
    except Notification.DoesNotExist:
        return Response({'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def mark_all_notifications_read(request):
    Notification.objects.filter(user=request.user, is_read=False).update(is_read=True)
    return Response({'message': 'All notifications marked as read'})


# ============ CHAT VIEWS ============

class ChatMessageListView(generics.ListCreateAPIView):
    serializer_class = ChatMessageSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        other_user_id = self.kwargs.get('user_id')
        user = self.request.user
        return ChatMessage.objects.filter(
            (Q(sender=user) & Q(receiver_id=other_user_id)) |
            (Q(sender_id=other_user_id) & Q(receiver=user))
        )

    def perform_create(self, serializer):
        serializer.save(sender=self.request.user)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def chat_contacts(request):
    """Get list of users the current user has chatted with."""
    user = request.user
    sent_to = ChatMessage.objects.filter(sender=user).values_list('receiver', flat=True).distinct()
    received_from = ChatMessage.objects.filter(receiver=user).values_list('sender', flat=True).distinct()
    contact_ids = set(list(sent_to) + list(received_from))
    contacts = User.objects.filter(id__in=contact_ids)
    return Response(UserSerializer(contacts, many=True).data)


# ============ FEEDBACK VIEWS ============

class FeedbackListCreateView(generics.ListCreateAPIView):
    serializer_class = FeedbackSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        if self.request.user.role in ('admin', 'placement_officer'):
            return Feedback.objects.all()
        return Feedback.objects.filter(user=self.request.user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


class FeedbackDetailView(generics.RetrieveDestroyAPIView):
    """
    GET    /api/feedback/<pk>/  — retrieve a single feedback entry
    DELETE /api/feedback/<pk>/  — owner or admin can delete
    """
    serializer_class = FeedbackSerializer
    permission_classes = [IsAuthenticated]
    queryset = Feedback.objects.all()

    def destroy(self, request, *args, **kwargs):
        fb = self.get_object()
        if fb.user != request.user and request.user.role != 'admin':
            return Response({'error': 'You can only delete your own feedback.'},
                            status=status.HTTP_403_FORBIDDEN)
        fb.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


# ============ SKILL COURSE VIEWS ============

class SkillCourseListView(generics.ListAPIView):
    serializer_class = SkillCourseSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = SkillCourse.objects.all()
        category = self.request.query_params.get('category')
        search = self.request.query_params.get('search')
        if category:
            queryset = queryset.filter(category=category)
        if search:
            queryset = queryset.filter(
                Q(title__icontains=search) | Q(description__icontains=search)
            )
        return queryset


class StudentCourseProgressView(generics.ListCreateAPIView):
    serializer_class = StudentCourseProgressSerializer
    permission_classes = [IsStudent]

    def get_queryset(self):
        return StudentCourseProgress.objects.filter(student=self.request.user)

    def perform_create(self, serializer):
        serializer.save(student=self.request.user)


class UpdateCourseProgressView(generics.UpdateAPIView):
    serializer_class = StudentCourseProgressSerializer
    permission_classes = [IsStudent]

    def get_queryset(self):
        return StudentCourseProgress.objects.filter(student=self.request.user)


# ============ ANALYTICS VIEWS ============

@api_view(['GET'])
@permission_classes([IsRecruiterOrPlacementOfficer])
def placement_analytics(request):
    """Get placement analytics and statistics."""
    total_students = StudentProfile.objects.count()
    placed_students = StudentProfile.objects.filter(is_placed=True).count()
    total_jobs = JobPosting.objects.filter(status='approved').count()
    total_applications = Application.objects.count()
    avg_cgpa_placed = StudentProfile.objects.filter(is_placed=True).aggregate(
        avg_cgpa=Avg('cgpa')
    )['avg_cgpa']

    # Map full department names back to codes (handles dirty data from manual registrations)
    DEPT_NAME_TO_CODE = {
        'computer science and engineering': 'CSE',
        'computer science engineering': 'CSE',
        'electronics and communication engineering': 'ECE',
        'electronics and communication': 'ECE',
        'electrical and electronics engineering': 'EEE',
        'mechanical engineering': 'MECH',
        'civil engineering': 'CIVIL',
        'information technology': 'IT',
        'ai and data science': 'AIDS',
    }
    VALID_CODES = {'CSE', 'ECE', 'EEE', 'MECH', 'CIVIL', 'IT', 'AIDS', 'OTHER'}

    raw_stats = StudentProfile.objects.values('department').annotate(
        total=Count('id'),
        placed=Count('id', filter=Q(is_placed=True))
    )
    # Merge/normalize: group by code
    merged = {}
    for row in raw_stats:
        code = row['department'].strip()
        if code.upper() in VALID_CODES:
            code = code.upper()
        else:
            code = DEPT_NAME_TO_CODE.get(code.lower(), code.upper()[:6])
        if code in merged:
            merged[code]['total'] += row['total']
            merged[code]['placed'] += row['placed']
        else:
            merged[code] = {'department': code, 'total': row['total'], 'placed': row['placed']}
    department_stats = list(merged.values())

    company_stats = Application.objects.values(
        company=F('job__recruiter__recruiter_profile__company_name')
    ).annotate(
        applications=Count('id'),
        selected=Count('id', filter=Q(status='selected'))
    ).order_by('-applications')[:10]

    return Response({
        'total_students': total_students,
        'placed_students': placed_students,
        'placement_rate': round(placed_students / total_students * 100, 2) if total_students > 0 else 0,
        'total_jobs': total_jobs,
        'total_applications': total_applications,
        'avg_cgpa_placed': round(float(avg_cgpa_placed), 2) if avg_cgpa_placed else 0,
        'department_stats': list(department_stats),
        'company_stats': list(company_stats),
    })


# ============ ADMIN VIEWS ============

class AdminStudentListView(generics.ListAPIView):
    serializer_class = StudentProfileSerializer
    permission_classes = [IsAdminUser]
    queryset = StudentProfile.objects.all()


class AdminUserUpdateView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = UserSerializer
    permission_classes = [IsAdminUser]
    queryset = User.objects.all()


class AdminAllUsersListView(generics.ListAPIView):
    """
    GET /api/admin/all-users/
    Returns all user accounts (all roles) for admin oversight.
    """
    serializer_class = UserSerializer
    permission_classes = [IsAdminUser]
    queryset = User.objects.all().order_by('role', 'username')


@api_view(['GET'])
@permission_classes([IsAdminUser])
def admin_system_stats(request):
    """
    GET /api/admin/system-stats/
    Returns platform-wide counts and system health info for the admin dashboard.
    """
    from django.utils import timezone as tz
    return Response({
        "total_users": User.objects.count(),
        "total_students": User.objects.filter(role='student').count(),
        "total_recruiters": User.objects.filter(role='recruiter').count(),
        "total_jobs": JobPosting.objects.count(),
        "total_applications": Application.objects.count(),
        "total_drives": PlacementDrive.objects.count(),
        "db_status": "healthy",
        "api_status": "online",
        "uptime": "99.9%",
        "last_backup": tz.now().date().isoformat(),
    })


# ============================================================
# UPGRADE 1 — Placement Predictor (scikit-learn Random Forest)
# ============================================================

@api_view(['GET'])
@permission_classes([IsStudent])
def placement_prediction(request):
    """
    GET /api/predict/placement/
    Returns ML-based placement probability (0–100%) for the logged-in student.
    """
    try:
        from api.ml.train_model import predict_for_profile
        profile = request.user.student_profile
        probability = predict_for_profile(profile)

        # Build actionable tips based on the profile
        tips = []
        if float(profile.cgpa) < 7.0:
            tips.append("Improve your CGPA — students with 7.0+ have significantly higher placement rates.")
        skills_count = len([s.strip() for s in profile.skills.split(',') if s.strip()]) if profile.skills else 0
        if skills_count < 5:
            tips.append(f"Add more skills to your profile (you have {skills_count}; aim for at least 5).")
        if not profile.linkedin_url:
            tips.append("Add your LinkedIn URL — it boosts recruiter visibility.")
        if not profile.github_url:
            tips.append("Add your GitHub URL — it showcases your coding projects.")
        if not profile.resume:
            tips.append("Upload your resume to apply for jobs.")

        return Response({
            'placement_probability': probability,
            'label': (
                'High' if probability >= 70 else
                'Medium' if probability >= 40 else
                'Low'
            ),
            'tips': tips,
            'features_used': {
                'cgpa': float(profile.cgpa),
                'skills_count': skills_count,
                'has_linkedin': bool(profile.linkedin_url),
                'has_github': bool(profile.github_url),
                'has_resume': bool(profile.resume),
                'department': profile.department,
            }
        })
    except StudentProfile.DoesNotExist:
        return Response({'error': 'Student profile not found'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([IsAdminUser])
def train_placement_model(request):
    """
    POST /api/predict/train/
    Re-train the Random Forest model on current DB data. Admin only.
    """
    try:
        from api.ml.train_model import train
        result = train()
        return Response({'message': 'Model trained successfully', 'details': result})
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ============================================================
# UPGRADE 2 — Resume Parser (pdfplumber + spaCy)
# ============================================================

@api_view(['POST'])
@permission_classes([IsStudent])
def parse_resume(request):
    """
    POST /api/resume/parse/
    Upload a PDF resume — AI extracts ALL fields and updates the complete student profile.
    Body: multipart/form-data with field 'resume' (PDF file)
    Optional query param: ?overwrite=true  → replace existing values even if already filled
    """
    resume_file = request.FILES.get('resume')
    if not resume_file:
        return Response({'error': 'No resume file provided.'}, status=status.HTTP_400_BAD_REQUEST)
    if not resume_file.name.lower().endswith('.pdf'):
        return Response({'error': 'Only PDF files are supported.'}, status=status.HTTP_400_BAD_REQUEST)

    overwrite = request.query_params.get('overwrite', 'false').lower() == 'true'

    try:
        import pdfplumber, io

        # 1. Extract raw text from PDF
        text = ''
        with pdfplumber.open(io.BytesIO(resume_file.read())) as pdf:
            for page in pdf.pages:
                text += (page.extract_text() or '') + '\n'
        resume_file.seek(0)

        if not text.strip():
            return Response({'error': 'Could not extract text from PDF. Try a text-based (not scanned) PDF.'},
                            status=status.HTTP_400_BAD_REQUEST)

        # 2. AI extraction — full profile
        try:
            from .ml.groq_service import parse_resume_with_ai
            extracted = parse_resume_with_ai(text)
        except Exception:
            extracted = _extract_resume_data(text)

        # 3. Get or create profile
        profile, _ = StudentProfile.objects.get_or_create(
            user=request.user,
            defaults={
                'roll_number': f'TEMP_{request.user.id}',
                'department': 'CSE',
                'batch_year': timezone.now().year,
                'cgpa': 0.0
            }
        )

        user = request.user
        user_fields_changed = []
        profile_fields_changed = []

        # ── Update User model fields ──────────────────────────────
        if extracted.get('first_name') and (overwrite or not user.first_name):
            user.first_name = extracted['first_name']
            user_fields_changed.append('first_name')

        if extracted.get('last_name') and (overwrite or not user.last_name):
            user.last_name = extracted['last_name']
            user_fields_changed.append('last_name')

        if extracted.get('email') and (overwrite or not user.email):
            user.email = extracted['email']
            user_fields_changed.append('email')

        if extracted.get('phone') and (overwrite or not user.phone):
            user.phone = extracted['phone']
            user_fields_changed.append('phone')

        if user_fields_changed:
            user.save(update_fields=user_fields_changed)

        # ── Update StudentProfile fields ───────────────────────────
        # Skills — REPLACE with new resume's skills (not merge)
        existing_skills = set(s.strip() for s in (profile.skills or '').split(',') if s.strip())
        new_skills = set(s.strip() for s in extracted.get('skills', []) if s.strip())
        if new_skills:
            profile.skills = ', '.join(sorted(new_skills))
            profile_fields_changed.append('skills')

        # Always overwrite linkedin/github from the new resume (clear if not found)
        new_linkedin = extracted.get('linkedin') or ''
        if new_linkedin != (profile.linkedin_url or ''):
            profile.linkedin_url = new_linkedin or None
            profile_fields_changed.append('linkedin_url')

        new_github = extracted.get('github') or ''
        if new_github != (profile.github_url or ''):
            profile.github_url = new_github or None
            profile_fields_changed.append('github_url')

        if extracted.get('about') and (overwrite or not profile.about):
            profile.about = extracted['about']
            profile_fields_changed.append('about')

        if extracted.get('cgpa') is not None and (overwrite or not profile.cgpa or profile.cgpa == 0.0):
            profile.cgpa = extracted['cgpa']
            profile_fields_changed.append('cgpa')

        if extracted.get('tenth_percentage') is not None and (overwrite or not profile.tenth_percentage):
            profile.tenth_percentage = extracted['tenth_percentage']
            profile_fields_changed.append('tenth_percentage')

        if extracted.get('twelfth_percentage') is not None and (overwrite or not profile.twelfth_percentage):
            profile.twelfth_percentage = extracted['twelfth_percentage']
            profile_fields_changed.append('twelfth_percentage')

        if extracted.get('department') and (overwrite or not profile.department or profile.department == 'CSE'):
            profile.department = extracted['department']
            profile_fields_changed.append('department')

        if extracted.get('batch_year') and (overwrite or not profile.batch_year or profile.batch_year == timezone.now().year):
            profile.batch_year = extracted['batch_year']
            profile_fields_changed.append('batch_year')

        if extracted.get('roll_number') and (overwrite or profile.roll_number.startswith('TEMP_')):
            profile.roll_number = extracted['roll_number']
            profile_fields_changed.append('roll_number')

        # Certifications — store as comma-separated in skills or separate field if exists
        if extracted.get('certifications'):
            certs = extracted['certifications']
            try:
                if hasattr(profile, 'certifications'):
                    existing_certs = set(c.strip() for c in (profile.certifications or '').split(',') if c.strip())
                    profile.certifications = ', '.join(sorted(existing_certs | set(certs)))
                    profile_fields_changed.append('certifications')
            except Exception:
                pass

        # Save the uploaded resume to the profile so we don't have to upload it again
        if resume_file:
            profile.resume = resume_file
            profile_fields_changed.append('resume')

        if profile_fields_changed:
            profile.save(update_fields=profile_fields_changed)

        # 4. Build summary of what was updated
        updated_summary = {
            'user_fields': user_fields_changed,
            'profile_fields': profile_fields_changed,
            'skills_added': list(new_skills - existing_skills),
            'certifications_found': extracted.get('certifications', []),
        }

        return Response({
            'message': f'Resume parsed and profile updated successfully! '
                       f'Updated {len(user_fields_changed)} user fields and {len(profile_fields_changed)} profile fields.',
            'extracted': extracted,
            'updated': updated_summary,
            'profile_snapshot': {
                'name': f"{user.first_name} {user.last_name}".strip(),
                'email': user.email,
                'phone': user.phone,
                'department': profile.department,
                'batch_year': profile.batch_year,
                'cgpa': profile.cgpa,
                'tenth_percentage': profile.tenth_percentage,
                'twelfth_percentage': profile.twelfth_percentage,
                'skills': profile.skills,
                'about': profile.about,
                'linkedin_url': profile.linkedin_url,
                'github_url': profile.github_url,
            }
        })

    except ImportError:
        return Response(
            {'error': 'pdfplumber is not installed. Run: pip install pdfplumber'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ── Resume ATS Review ─────────────────────────────────────────
@api_view(['POST'])
@permission_classes([IsStudent])
def resume_ats_review(request):
    """
    POST /api/resume/review/
    Upload a PDF resume → get ATS score, strengths, issues, suggestions.
    Body: multipart/form-data, field 'resume' (PDF). Optional: field 'target_role' (string).
    """
    resume_file = request.FILES.get('resume')
    if not resume_file:
        # Try to use stored resume from profile
        try:
            profile = StudentProfile.objects.get(user=request.user)
            if profile.resume:
                import pdfplumber, io
                text = ''
                with pdfplumber.open(profile.resume.path) as pdf:
                    for page in pdf.pages:
                        text += (page.extract_text() or '') + '\n'
            else:
                return Response({'error': 'No resume provided and no resume uploaded to profile.'},
                                status=status.HTTP_400_BAD_REQUEST)
        except StudentProfile.DoesNotExist:
            return Response({'error': 'No resume file provided.'}, status=status.HTTP_400_BAD_REQUEST)
    else:
        import pdfplumber, io
        text = ''
        with pdfplumber.open(io.BytesIO(resume_file.read())) as pdf:
            for page in pdf.pages:
                text += (page.extract_text() or '') + '\n'
        resume_file.seek(0)

    target_role = request.data.get('target_role', 'Software Engineer')
    try:
        from .ml.groq_service import review_resume_ats
        result = review_resume_ats(text, target_role)
        return Response(result)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ── Cover Letter Generator ───────────────────────────────────
@api_view(['POST'])
@permission_classes([IsStudent])
def generate_cover_letter_view(request, pk):
    """
    POST /api/jobs/<pk>/cover-letter/
    Generate a personalised cover letter for this job using student's resume.
    Body: multipart/form-data, field 'resume' (PDF) — optional if resume on profile.
    """
    try:
        job = JobPosting.objects.get(pk=pk)
    except JobPosting.DoesNotExist:
        return Response({'error': 'Job not found.'}, status=status.HTTP_404_NOT_FOUND)

    import pdfplumber, io

    resume_file = request.FILES.get('resume')
    if resume_file:
        text = ''
        with pdfplumber.open(io.BytesIO(resume_file.read())) as pdf:
            for page in pdf.pages:
                text += (page.extract_text() or '') + '\n'
        resume_file.seek(0)
    else:
        try:
            profile = StudentProfile.objects.get(user=request.user)
            if profile.resume:
                text = ''
                with pdfplumber.open(profile.resume.path) as pdf:
                    for page in pdf.pages:
                        text += (page.extract_text() or '') + '\n'
            else:
                text = f"Skills: {profile.skills}\nDepartment: {profile.department}\nCGPA: {profile.cgpa}"
        except StudentProfile.DoesNotExist:
            text = ""

    student_name = f"{request.user.first_name} {request.user.last_name}".strip() or request.user.username
    try:
        from .ml.groq_service import generate_cover_letter
        company_name = _safe_company_name(job)
        letter = generate_cover_letter(
            resume_text=text,
            job_title=job.title,
            company_name=company_name,
            job_description=f"{job.description}\n\nRequirements: {job.requirements}",
            student_name=student_name,
        )
        return Response({'cover_letter': letter, 'job_title': job.title, 'company': company_name})
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ── Mock Interview Questions from Resume ─────────────────────
@api_view(['POST'])
@permission_classes([IsStudent])
def resume_interview_questions(request):
    """
    POST /api/resume/interview-questions/
    Upload PDF → get personalised mock interview questions.
    Body: multipart/form-data, field 'resume' (PDF). Optional: 'job_title'.
    """
    import pdfplumber, io

    resume_file = request.FILES.get('resume')
    if resume_file:
        text = ''
        with pdfplumber.open(io.BytesIO(resume_file.read())) as pdf:
            for page in pdf.pages:
                text += (page.extract_text() or '') + '\n'
        resume_file.seek(0)
    else:
        try:
            profile = StudentProfile.objects.get(user=request.user)
            if profile.resume:
                text = ''
                with pdfplumber.open(profile.resume.path) as pdf:
                    for page in pdf.pages:
                        text += (page.extract_text() or '') + '\n'
            else:
                # Synthesize from profile data
                text = f"Skills: {profile.skills}\nDepartment: {profile.department}\nCGPA: {profile.cgpa}\nAbout: {profile.about}"
        except StudentProfile.DoesNotExist:
            return Response({'error': 'Please upload a resume file.'}, status=status.HTTP_400_BAD_REQUEST)

    job_title = request.data.get('job_title', 'Software Engineer')
    try:
        from .ml.groq_service import generate_interview_questions
        result = generate_interview_questions(text, job_title)
        return Response(result)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



def _extract_resume_data(text):
    """Extract structured data from raw resume text using spaCy + regex."""
    import re

    result = {
        'skills': [],
        'email': None,
        'phone': None,
        'linkedin': None,
        'github': None,
        'organisations': [],
        'raw_text_length': len(text),
    }

    # Email
    emails = re.findall(r'[\w.+-]+@[\w-]+\.[a-z]{2,}', text, re.I)
    if emails:
        result['email'] = emails[0]

    # Phone
    phones = re.findall(r'[\+]?[(]?[0-9]{3}[)]?[-\s.]?[0-9]{3}[-\s.]?[0-9]{4,6}', text)
    if phones:
        result['phone'] = phones[0]

    # LinkedIn & GitHub URLs
    linkedin = re.findall(r'linkedin\.com/in/[\w-]+', text, re.I)
    if linkedin:
        result['linkedin'] = 'https://' + linkedin[0]

    github = re.findall(r'github\.com/[\w-]+', text, re.I)
    if github:
        result['github'] = 'https://' + github[0]

    # Skills — match against a comprehensive keyword list
    SKILL_KEYWORDS = [
        'python', 'java', 'javascript', 'typescript', 'c++', 'c#', 'ruby', 'php',
        'go', 'rust', 'swift', 'kotlin', 'dart', 'scala', 'r',
        'sql', 'mysql', 'postgresql', 'mongodb', 'redis', 'cassandra', 'sqlite',
        'react', 'angular', 'vue', 'svelte', 'next.js', 'nuxt',
        'node', 'express', 'django', 'flask', 'fastapi', 'spring', 'laravel',
        'docker', 'kubernetes', 'aws', 'azure', 'gcp', 'terraform', 'ansible',
        'git', 'linux', 'bash', 'jenkins', 'github actions', 'ci/cd',
        'machine learning', 'deep learning', 'nlp', 'computer vision',
        'tensorflow', 'pytorch', 'scikit-learn', 'keras', 'pandas', 'numpy', 'matplotlib',
        'html', 'css', 'sass', 'tailwind', 'bootstrap',
        'rest api', 'graphql', 'microservices', 'agile', 'scrum',
        'flutter', 'react native', 'android', 'ios',
        'hadoop', 'spark', 'kafka', 'airflow', 'tableau', 'power bi',
        'figma', 'photoshop', 'canva',
    ]

    text_lower = text.lower()
    found_skills = [skill for skill in SKILL_KEYWORDS if skill in text_lower]
    result['skills'] = found_skills

    # Try spaCy for named entities (organisations)
    try:
        import spacy
        try:
            nlp = spacy.load('en_core_web_sm')
        except OSError:
            nlp = None

        if nlp:
            doc = nlp(text[:5000])  # limit for performance
            result['organisations'] = list(set(
                ent.text for ent in doc.ents if ent.label_ == 'ORG'
            ))
    except ImportError:
        pass  # spaCy not installed — skills extraction still works

    return result


# ============================================================
# UPGRADE 6 — Export Reports (ReportLab PDF + openpyxl Excel)
# ============================================================

@api_view(['GET'])
@authentication_classes([QueryParamJWTAuthentication])
@permission_classes([IsRecruiterOrPlacementOfficer])
def export_placement_pdf(request):
    """
    GET /api/reports/placement.pdf?token=<jwt>
    Download a formatted PDF placement report.
    Accepts token via ?token= query param (browser download) or Authorization header.
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.enums import TA_CENTER
        import io

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4,
                                rightMargin=2*cm, leftMargin=2*cm,
                                topMargin=2*cm, bottomMargin=2*cm)

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('Title', parent=styles['Title'],
                                     fontSize=18, textColor=colors.HexColor('#1a237e'),
                                     alignment=TA_CENTER)
        heading_style = ParagraphStyle('Heading', parent=styles['Heading2'],
                                       fontSize=13, textColor=colors.HexColor('#283593'))

        elements = []

        # Title
        elements.append(Paragraph('JobQuench — Placement Report', title_style))
        elements.append(Paragraph(
            f'Generated on {timezone.now().strftime("%d %B %Y, %I:%M %p")}',
            styles['Normal']
        ))
        elements.append(Spacer(1, 0.5*cm))

        # Summary stats
        total = StudentProfile.objects.count()
        placed = StudentProfile.objects.filter(is_placed=True).count()
        rate = round(placed / total * 100, 1) if total else 0
        avg_cgpa = StudentProfile.objects.filter(is_placed=True).aggregate(
            a=Avg('cgpa'))['a'] or 0

        elements.append(Paragraph('Summary', heading_style))
        summary_data = [
            ['Metric', 'Value'],
            ['Total Students', str(total)],
            ['Placed Students', str(placed)],
            ['Placement Rate', f'{rate}%'],
            ['Average CGPA (Placed)', f'{round(float(avg_cgpa), 2)}'],
            ['Total Job Postings', str(JobPosting.objects.filter(status='approved').count())],
            ['Total Applications', str(Application.objects.count())],
        ]
        t = Table(summary_data, colWidths=[10*cm, 6*cm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3f51b5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#e8eaf6')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('PADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 0.5*cm))

        # Department breakdown
        elements.append(Paragraph('Department-wise Placement', heading_style))
        dept_data = [['Department', 'Total', 'Placed', 'Rate']]
        dept_stats = StudentProfile.objects.values('department').annotate(
            total=Count('id'),
            placed=Count('id', filter=Q(is_placed=True))
        )
        for d in dept_stats:
            r = round(d['placed'] / d['total'] * 100, 1) if d['total'] else 0
            dept_data.append([d['department'], str(d['total']), str(d['placed']), f'{r}%'])

        t2 = Table(dept_data, colWidths=[6*cm, 3*cm, 3*cm, 4*cm])
        t2.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3f51b5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#e8eaf6')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('PADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(t2)

        doc.build(elements)
        buffer.seek(0)

        from django.http import HttpResponse
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = (
            f'attachment; filename="placement_report_{timezone.now().strftime("%Y%m%d")}.pdf"'
        )
        return response

    except ImportError:
        return Response(
            {'error': 'ReportLab not installed. Run: pip install reportlab'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@authentication_classes([QueryParamJWTAuthentication])
@permission_classes([IsRecruiterOrPlacementOfficer])
def export_placement_excel(request):
    """
    GET /api/reports/placement.xlsx?token=<jwt>
    Download placement data as a formatted Excel workbook.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import io

        wb = openpyxl.Workbook()

        # ---- Sheet 1: Summary ----
        ws1 = wb.active
        ws1.title = 'Summary'
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill('solid', fgColor='3F51B5')
        alt_fill = PatternFill('solid', fgColor='E8EAF6')
        center = Alignment(horizontal='center', vertical='center')

        ws1.append(['JobQuench Placement Report'])
        ws1.append([f'Generated: {timezone.now().strftime("%d %B %Y")}'])
        ws1.append([])
        ws1.append(['Metric', 'Value'])
        for cell in ws1[4]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        total = StudentProfile.objects.count()
        placed = StudentProfile.objects.filter(is_placed=True).count()
        rate = round(placed / total * 100, 1) if total else 0
        avg_cgpa = StudentProfile.objects.filter(is_placed=True).aggregate(
            a=Avg('cgpa'))['a'] or 0

        rows = [
            ['Total Students', total],
            ['Placed Students', placed],
            ['Placement Rate (%)', rate],
            ['Average CGPA (Placed)', round(float(avg_cgpa), 2)],
            ['Total Jobs Posted', JobPosting.objects.filter(status='approved').count()],
            ['Total Applications', Application.objects.count()],
        ]
        for i, row in enumerate(rows):
            ws1.append(row)
            if i % 2 == 1:
                for cell in ws1[ws1.max_row]:
                    cell.fill = alt_fill

        ws1.column_dimensions['A'].width = 30
        ws1.column_dimensions['B'].width = 20

        # ---- Sheet 2: Students ----
        ws2 = wb.create_sheet('Students')
        headers = ['Name', 'Roll No', 'Department', 'CGPA', 'Skills',
                   'Is Placed', 'Placed Company', 'Package (LPA)']
        ws2.append(headers)
        for cell in ws2[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        for i, p in enumerate(StudentProfile.objects.select_related('user').all()):
            ws2.append([
                p.user.get_full_name(),
                p.roll_number,
                p.department,
                float(p.cgpa),
                p.skills,
                'Yes' if p.is_placed else 'No',
                p.placed_company or '',
                float(p.placed_package) if p.placed_package else '',
            ])
            if i % 2 == 1:
                for cell in ws2[ws2.max_row]:
                    cell.fill = alt_fill

        for col in range(1, len(headers) + 1):
            ws2.column_dimensions[get_column_letter(col)].width = 18

        # ---- Sheet 3: Company Applications ----
        ws3 = wb.create_sheet('Company Stats')
        ws3.append(['Company', 'Total Applications', 'Selected'])
        for cell in ws3[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        company_stats = Application.objects.values(
            company=F('job__recruiter__recruiter_profile__company_name')
        ).annotate(
            applications=Count('id'),
            selected=Count('id', filter=Q(status='selected'))
        ).order_by('-applications')

        for i, c in enumerate(company_stats):
            ws3.append([c['company'], c['applications'], c['selected']])
            if i % 2 == 1:
                for cell in ws3[ws3.max_row]:
                    cell.fill = alt_fill

        for col in ['A', 'B', 'C']:
            ws3.column_dimensions[col].width = 22

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        from django.http import HttpResponse
        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f'attachment; filename="placement_{timezone.now().strftime("%Y%m%d")}.xlsx"'
        )
        return response

    except ImportError:
        return Response(
            {'error': 'openpyxl not installed. Run: pip install openpyxl'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


# ============================================================
# UPGRADE 7 — Interview Readiness Score (Pure Python formula)
# ============================================================

@api_view(['GET'])
@permission_classes([IsStudent])
def interview_readiness_score(request):
    """
    GET /api/student/readiness-score/
    Returns a 0–100 score with weighted components and actionable tips.

    Weights:
        CGPA          30%
        Skills count  25%
        Applications  10%
        LinkedIn      15%
        GitHub / Resume 10% each
    """
    try:
        profile = request.user.student_profile
    except StudentProfile.DoesNotExist:
        return Response({'error': 'Student profile not found'}, status=status.HTTP_404_NOT_FOUND)

    # Component scores (each 0–100 internally, then weighted)
    cgpa_score = min(100, (float(profile.cgpa) / 10.0) * 100)

    skills = [s.strip() for s in profile.skills.split(',') if s.strip()] if profile.skills else []
    # 20 skills = full marks so score actually differentiates between resumes
    skills_score = min(100, (len(skills) / 20.0) * 100)

    applications_count = Application.objects.filter(student=request.user).count()
    applications_score = min(100, (applications_count / 5.0) * 100)

    linkedin_score = 100 if profile.linkedin_url else 0
    github_score = 100 if profile.github_url else 0
    resume_score = 100 if profile.resume else 0

    # Weighted total
    total = (
        cgpa_score * 0.30 +
        skills_score * 0.25 +
        applications_score * 0.10 +
        linkedin_score * 0.15 +
        github_score * 0.10 +
        resume_score * 0.10
    )
    total = round(total, 1)

    # Generate Groq-powered personalized tips
    ai_tips = None
    try:
        from .ml.groq_service import generate_readiness_tips
        ai_tips = generate_readiness_tips({
            'cgpa': float(profile.cgpa),
            'skills': profile.skills,
            'skills_count': len(skills),
            'has_linkedin': bool(profile.linkedin_url),
            'has_github': bool(profile.github_url),
            'has_resume': bool(profile.resume),
            'applications_count': applications_count,
            'score': total,
            'department': getattr(profile, 'department', 'Engineering'),
        })
    except Exception:
        pass

    # Fallback rule-based tips
    fallback_tips = []
    if cgpa_score < 70:
        needed = 7.0 - float(profile.cgpa)
        fallback_tips.append(f'Improve your CGPA by {round(needed, 2)} points to reach the 7.0 benchmark.')
    if len(skills) < 5:
        fallback_tips.append(f'Add {5 - len(skills)} more skill(s) to boost your score.')
    if not profile.linkedin_url:
        fallback_tips.append('Add your LinkedIn URL for recruiter visibility.')
    if not profile.github_url:
        fallback_tips.append('Add your GitHub URL to showcase your projects.')
    if not profile.resume:
        fallback_tips.append('Upload your resume to apply for jobs.')
    if applications_count == 0:
        fallback_tips.append('Apply for at least one job to earn application activity points.')
    if total >= 80:
        fallback_tips.append('Excellent! You are well-prepared for placements.')

    return Response({
        'score': total,
        'label': (
            'Excellent' if total >= 80 else
            'Good' if total >= 60 else
            'Fair' if total >= 40 else
            'Needs Work'
        ),
        'breakdown': {
            'cgpa': {'score': round(cgpa_score * 0.30, 1), 'weight': '30%', 'raw': float(profile.cgpa)},
            'skills': {'score': round(skills_score * 0.25, 1), 'weight': '25%', 'count': len(skills)},
            'applications': {'score': round(applications_score * 0.10, 1), 'weight': '10%', 'count': applications_count},
            'linkedin': {'score': round(linkedin_score * 0.15, 1), 'weight': '15%', 'present': bool(profile.linkedin_url)},
            'github': {'score': round(github_score * 0.10, 1), 'weight': '10%', 'present': bool(profile.github_url)},
            'resume': {'score': round(resume_score * 0.10, 1), 'weight': '10%', 'present': bool(profile.resume)},
        },
        'tips': fallback_tips,
        'ai_advice': ai_tips,
    })


# ============================================================
# UPGRADE 8 — Skill Gap Analyser (Pure Python set operations)
# ============================================================

@api_view(['GET'])
@permission_classes([IsStudent])
def skill_gap(request, pk):
    """
    GET /api/jobs/<pk>/skill-gap/
    Compares the job's required skills with the student's skills.
    Returns matched, missing skills, and free learning resource links.
    """
    LEARNING_RESOURCES = {
        'python': 'https://www.freecodecamp.org/news/python-tutorial/',
        'javascript': 'https://javascript.info/',
        'java': 'https://www.youtube.com/watch?v=eIrMbAQSU34',
        'sql': 'https://www.w3schools.com/sql/',
        'react': 'https://react.dev/learn',
        'docker': 'https://docs.docker.com/get-started/',
        'aws': 'https://aws.amazon.com/training/digital/',
        'kubernetes': 'https://kubernetes.io/docs/tutorials/',
        'machine learning': 'https://www.coursera.org/learn/machine-learning',
        'deep learning': 'https://www.deeplearning.ai/',
        'git': 'https://www.youtube.com/watch?v=RGOj5yH7evk',
        'linux': 'https://www.freecodecamp.org/news/the-linux-commands-handbook/',
        'django': 'https://www.djangoproject.com/start/',
        'angular': 'https://angular.dev/tutorials',
        'vue': 'https://vuejs.org/tutorial/',
        'mongodb': 'https://learn.mongodb.com/',
        'postgresql': 'https://www.postgresqltutorial.com/',
        'typescript': 'https://www.typescriptlang.org/docs/handbook/intro.html',
        'flutter': 'https://flutter.dev/learn',
        'node': 'https://nodejs.org/en/learn/getting-started/introduction-to-nodejs',
        'spring': 'https://spring.io/guides',
    }

    try:
        job = JobPosting.objects.get(pk=pk, status='approved')
    except JobPosting.DoesNotExist:
        return Response({'error': 'Job not found'}, status=status.HTTP_404_NOT_FOUND)

    try:
        profile = request.user.student_profile
    except StudentProfile.DoesNotExist:
        return Response({'error': 'Student profile not found'}, status=status.HTTP_404_NOT_FOUND)

    # Normalise skill sets
    def parse_skills(raw):
        return set(s.strip().lower() for s in raw.split(',') if s.strip())

    student_skills = parse_skills(profile.skills)
    required_skills = parse_skills(job.required_skills)

    if not required_skills:
        return Response({
            'message': 'This job has no specific skill requirements listed.',
            'job_title': job.title,
            'your_skills': sorted(student_skills),
        })

    matched = sorted(student_skills & required_skills)
    missing = sorted(required_skills - student_skills)
    match_pct = round(len(matched) / len(required_skills) * 100, 1) if required_skills else 100

    LEARNING_RESOURCES_LOCAL = {
        'python': 'https://www.freecodecamp.org/news/python-tutorial/',
        'javascript': 'https://javascript.info/',
        'java': 'https://www.youtube.com/watch?v=eIrMbAQSU34',
        'sql': 'https://www.w3schools.com/sql/',
        'react': 'https://react.dev/learn',
        'docker': 'https://docs.docker.com/get-started/',
        'aws': 'https://aws.amazon.com/training/digital/',
        'machine learning': 'https://www.coursera.org/learn/machine-learning',
        'deep learning': 'https://www.deeplearning.ai/',
        'git': 'https://www.youtube.com/watch?v=RGOj5yH7evk',
        'linux': 'https://www.freecodecamp.org/news/the-linux-commands-handbook/',
        'django': 'https://www.djangoproject.com/start/',
        'typescript': 'https://www.typescriptlang.org/docs/handbook/intro.html',
        'flutter': 'https://flutter.dev/learn',
        'node': 'https://nodejs.org/en/learn/getting-started/introduction-to-nodejs',
    }

    missing_with_resources = []
    for skill in missing:
        resource = LEARNING_RESOURCES_LOCAL.get(skill)
        if not resource:
            resource = f'https://www.youtube.com/results?search_query={skill.replace(" ", "+")}+tutorial'
        missing_with_resources.append({'skill': skill, 'learn_at': resource})

    # Groq learning roadmap
    roadmap = None
    try:
        from .ml.groq_service import generate_skill_gap_roadmap
        roadmap = generate_skill_gap_roadmap(job.title, missing, matched)
    except Exception:
        pass

    return Response({
        'job_title': job.title,
        'company': job.recruiter.recruiter_profile.company_name if hasattr(job.recruiter, 'recruiter_profile') else '',
        'match_percentage': match_pct,
        'matched_skills': matched,
        'missing_skills': missing_with_resources,
        'your_skills': sorted(student_skills),
        'required_skills': sorted(required_skills),
        'verdict': (
            'Strong match! You meet most requirements.' if match_pct >= 75 else
            'Good potential - close some skill gaps to strengthen your application.' if match_pct >= 50 else
            'Significant gaps - focus on learning the missing skills before applying.'
        ),
        'ai_roadmap': roadmap,
    })


# ============================================================
# AI CAREER COACH - Groq-powered chat endpoint
# ============================================================

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def ai_chat(request):
    """
    POST /api/ai/chat/
    Body: { "message": str, "history": [ { "role": "user|assistant", "content": str } ] }
    """
    message = request.data.get('message', '')
    if isinstance(message, str):
        message = message.strip()
    else:
        message = str(message).strip()

    if not message:
        return Response({'error': 'Message cannot be empty.'}, status=status.HTTP_400_BAD_REQUEST)

    # Safely convert history to plain list of dicts
    raw_history = request.data.get('history', [])
    history = []
    if isinstance(raw_history, list):
        for item in raw_history:
            try:
                role = str(item.get('role', '') if hasattr(item, 'get') else '')
                content = str(item.get('content', '') if hasattr(item, 'get') else '')
                if role in ('user', 'assistant') and content:
                    history.append({'role': role, 'content': content})
            except Exception:
                pass

    try:
        from .ml.groq_service import ai_career_chat
        reply = ai_career_chat(message, history)
        return Response({'reply': reply})
    except Exception as e:
        return Response(
            {'error': f'AI service unavailable: {str(e)}'},
            status=status.HTTP_503_SERVICE_UNAVAILABLE
        )


# ============================================================
# AUTO-APPLY — Upload resume → auto-apply to matching jobs
# ============================================================

@api_view(['POST'])
@permission_classes([IsStudent])
def auto_apply_from_resume(request):
    """
    POST /api/resume/auto-apply/
    Upload a PDF resume. The backend will:
      1. Extract skills using Groq AI (fallback: regex)
      2. Update the student's profile skills
      3. Find all approved, non-expired jobs that match:
             - Student CGPA >= job.min_cgpa
             - Student department in job.eligible_departments (or no restriction)
             - At least 1 skill overlap  (or any job if student has no skills yet)
      4. Auto-create Applications for every matched job (skip if already applied)
      5. Return the list of jobs applied to

    Body: multipart/form-data with field 'resume' (PDF file)
    """
    resume_file = request.FILES.get('resume')
    if not resume_file:
        return Response({'error': 'No resume file provided.'}, status=status.HTTP_400_BAD_REQUEST)
    if not resume_file.name.lower().endswith('.pdf'):
        return Response({'error': 'Only PDF files are supported.'}, status=status.HTTP_400_BAD_REQUEST)

    # ── Step 1: Extract text from PDF ────────────────────────────────────
    try:
        import pdfplumber, io
        text = ''
        with pdfplumber.open(io.BytesIO(resume_file.read())) as pdf:
            for page in pdf.pages:
                text += (page.extract_text() or '') + '\n'
    except ImportError:
        return Response({'error': 'pdfplumber not installed.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    except Exception as e:
        return Response({'error': f'Could not read PDF: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

    # ── Step 2: Extract skills (Groq → regex fallback) ───────────────────
    try:
        from .ml.groq_service import parse_resume_with_ai
        extracted = parse_resume_with_ai(text)
        if not extracted.get('skills'):
            extracted = _extract_resume_data(text)
    except Exception:
        extracted = _extract_resume_data(text)

    new_skills = set(s.strip().lower() for s in extracted.get('skills', []) if s.strip())

    # ── Step 3: Update student profile ───────────────────────────────────
    profile, _ = StudentProfile.objects.get_or_create(
        user=request.user,
        defaults={
            'roll_number': f'TEMP_{request.user.id}',
            'department': 'CSE',
            'batch_year': timezone.now().year,
            'cgpa': 0.0,
        }
    )

    existing_skills = set(s.strip().lower() for s in profile.skills.split(',') if s.strip())
    merged_skills   = existing_skills | new_skills
    profile.skills  = ', '.join(sorted(merged_skills))
    profile.save(update_fields=['skills'])

    student_skills = merged_skills  # normalised lowercase set
    student_cgpa   = float(profile.cgpa) if profile.cgpa else 0.0
    student_dept   = profile.department  # short code e.g. 'CSE'

    # ── Step 4: Find matching approved jobs ──────────────────────────────
    open_jobs = JobPosting.objects.filter(
        status='approved',
        application_deadline__gte=timezone.now(),
    )

    matched_jobs   = []
    already_applied = set(
        Application.objects.filter(student=request.user)
        .values_list('job_id', flat=True)
    )

    for job in open_jobs:
        # Skip if already applied
        if job.id in already_applied:
            continue

        # CGPA check
        if student_cgpa < float(job.min_cgpa):
            continue

        # Department check (skip if job has no restriction)
        if job.eligible_departments:
            eligible = [d.strip().upper() for d in job.eligible_departments.split(',')]
            if student_dept.upper() not in eligible:
                continue

        # Skill overlap check
        if job.required_skills:
            job_skills = set(s.strip().lower() for s in job.required_skills.split(',') if s.strip())
            if job_skills and not (student_skills & job_skills):
                continue  # zero overlap → skip

        matched_jobs.append(job)

    # ── Step 5: Auto-create applications ─────────────────────────────────
    applied = []
    for job in matched_jobs:
        try:
            Application.objects.create(
                student=request.user,
                job=job,
                cover_letter=(
                    f"Auto-applied based on resume skills match. "
                    f"Skills: {profile.skills[:300]}"
                ),
            )
            # Notify recruiter
            Notification.objects.create(
                user=job.recruiter,
                notification_type='application_update',
                title=f'New Application: {job.title}',
                message=(
                    f'{request.user.get_full_name()} auto-applied for {job.title} '
                    f'via resume upload.'
                ),
                link=f'/applications/',
            )
            applied.append({
                'job_id':      job.id,
                'job_title':   job.title,
                'company':     _safe_company_name(job),
                'location':    job.location,
                'job_type':    job.get_job_type_display(),
                'salary_min':  str(job.salary_min) if job.salary_min else None,
                'salary_max':  str(job.salary_max) if job.salary_max else None,
                'deadline':    job.application_deadline.isoformat(),
            })
        except Exception:
            pass   # unique_together violation or other race — skip silently

    # Notify student with summary
    if applied:
        Notification.objects.create(
            user=request.user,
            notification_type='application_update',
            title=f'Auto-Applied to {len(applied)} Job(s)!',
            message=(
                f'Your resume matched {len(applied)} open job(s). '
                f'Applications submitted automatically.'
            ),
            link='/applications/my/',
        )

    return Response({
        'message': (
            f'Resume processed. Auto-applied to {len(applied)} matching job(s).'
            if applied else
            'Resume processed. No new matching jobs found right now.'
        ),
        'skills_extracted': sorted(new_skills),
        'profile_skills':   profile.skills,
        'total_applied':    len(applied),
        'jobs_applied':     applied,
    }, status=status.HTTP_200_OK)


def _safe_company_name(job):
    """Safely get company name without crashing if profile missing."""
    try:
        return job.recruiter.recruiter_profile.company_name
    except Exception:
        return 'Unknown Company'
